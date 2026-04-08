"""Excel export with bold header row (openpyxl)."""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font


def build_workbook_response(filename, sheet_title, headers, rows):
    """
    filename: e.g. 'report.xlsx'
    sheet_title: max 31 chars for Excel sheet name
    headers: list of str
    rows: list of list of cell values
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Report")[:31]

    bold = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, filename
